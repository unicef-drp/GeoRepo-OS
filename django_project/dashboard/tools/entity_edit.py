import logging
import os
import csv
import re
import traceback
import openpyxl
from django.conf import settings
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.db.models import Max
from django.db.models.fields.files import FieldFile


from georepo.models.id_type import IdType
from georepo.models.language import Language
from georepo.models.base_task_request import PROCESSING, DONE, ERROR, PENDING
from georepo.models.entity import (
    GeographicalEntity,
    EntityId,
    EntityName
)
from dashboard.models.batch_edit import BatchEntityEdit, EntityEditResult
from georepo.utils.unique_code import parse_unique_code
from dashboard.models.notification import (
    Notification,
    NOTIF_TYPE_BATCH_ENTITY_EDIT
)
from georepo.tasks.dataset_view import check_affected_dataset_views


logger = logging.getLogger(__name__)
TEMP_OUTPUT_DIRECTORY = getattr(
    settings, 'FILE_UPLOAD_TEMP_DIR',
    '/home/web/media/tmp/batch_entity_edit')
TEMP_OUTPUT_DIRECTORY = (
    TEMP_OUTPUT_DIRECTORY if
    TEMP_OUTPUT_DIRECTORY is not None else
    '/home/web/media/tmp/batch_entity_edit'
)


if not os.path.exists(TEMP_OUTPUT_DIRECTORY):
    os.makedirs(TEMP_OUTPUT_DIRECTORY)


def try_delete_uploaded_file(file: FieldFile):
    try:
        file.delete(save=False)
    except Exception:
        logger.error('Failed to delete file!')


def format_entities_count(count):
    if count <= 1:
        return f'{count} entity'
    return f'{count} entities'


class BatchEntityEditBaseImporter(object):
    preview = False
    request = BatchEntityEdit.objects.none()
    headers = []
    total_rows = 0
    ucode_index = -1
    # store name/id fields with its index
    name_fields = {}
    id_fields = {}
    # validation of missing columns
    missing_columns = []
    # metadata
    id_types = IdType.objects.all()
    languages = Language.objects.all()
    # outputs
    output_headers = []
    csv_output_file_path = None
    csv_output_file = None
    csv_output_writer = None
    entity_ids = []
    entity_results = []

    def __init__(self, request: BatchEntityEdit, preview: bool) -> None:
        self.request = request
        self.preview = preview

    def process_started(self):
        # reset state, clear exiting output
        self.headers = []
        self.total_rows = 0
        self.ucode_index = -1
        self.name_fields = {}
        self.id_fields = {}
        self.missing_columns = []
        self.output_headers = []
        self.request.status = PROCESSING
        self.request.started_at = timezone.now()
        self.request.finished_at = None
        self.request.progress = 0
        self.request.errors = None
        if self.request.output_file:
            try_delete_uploaded_file(self.request.output_file)
            self.request.output_file = None
        if self.preview:
            if self.request.preview_file:
                try_delete_uploaded_file(self.request.preview_file)
                self.request.preview_file = None
        self.request.error_notes = None
        self.request.error_count = 0
        self.request.success_notes = None
        self.request.success_count = 0
        self.request.save(update_fields=[
            'status', 'started_at', 'progress', 'errors',
            'finished_at', 'output_file', 'error_notes',
            'error_count', 'success_notes',
            'success_count', 'preview_file'
        ])
        # reload metadata
        self.id_types = IdType.objects.all()
        self.languages = Language.objects.all()
        self.entity_ids = []
        self.entity_results = []
        # delete existing entity results
        if self.preview:
            entity_result_qs = EntityEditResult.objects.filter(
                batch_edit=self.request
            )
            entity_result_qs._raw_delete(entity_result_qs.db)

    def remove_csv_output_file(self):
        """Remove temporary output file."""
        if self.csv_output_file_path:
            try:
                if os.path.exists(self.csv_output_file_path):
                    os.remove(self.csv_output_file_path)
            except Exception as ex:
                logger.warn('Failed to delete csv output file: ')
                logger.warn(ex)
                logger.warn(traceback.format_exc())

    def process_ended(self, is_success, errors = None):
        # set final state, clear temp resource
        final_state = DONE if not self.preview else PENDING
        self.request.status = final_state if is_success else ERROR
        if not is_success:
            self.request.errors = errors
        self.request.finished_at = timezone.now()
        self.request.task_id = None
        self.request.save(update_fields=[
            'status', 'finished_at', 'errors', 'task_id'])
        if self.csv_output_file:
            self.csv_output_file.close()
            # open the file in binary to upload to blob storage
            with open(self.csv_output_file_path, 'rb') as output_file:
                if self.preview:
                    self.request.preview_file.save(
                        os.path.basename(output_file.name),
                        output_file
                    )
                else:
                    self.request.output_file.save(
                        os.path.basename(output_file.name),
                        output_file
                    )
            # delete the output
            self.remove_csv_output_file()
        if is_success and not self.preview and self.entity_ids:
            # trigger check_affected_dataset_views
            check_affected_dataset_views(
                self.request.dataset.id, self.entity_ids, [], False)
            # update task to 100
            self.request.progress = 100
            self.request.save(update_fields=['progress'])
        if not self.preview:
            # trigger notifications
            dataset = self.request.dataset
            message = (
                'Your batch entity edit for '
                f'{dataset.label}'
                ' has finished! Click here to view!'
            )
            payload = {
                'session': self.request.id,
                'dataset': dataset.id,
                'step': 2,
                'severity': 'success',
                'module': 'admin_boundaries'
            }
            Notification.objects.create(
                type=NOTIF_TYPE_BATCH_ENTITY_EDIT,
                message=message,
                recipient=self.request.submitted_by,
                payload=payload
            )

    def on_update_progress(self, row, total_count):
        max_progress = 100 if self.preview else 80
        # save progress of task
        self.request.progress = (
            (row / total_count) * max_progress if total_count else 0
        )
        self.request.save(update_fields=['progress'])

    def start(self):
        self.process_started()
        is_success = False
        error = None
        try:
            self.read_headers()
            self.process_headers()
            is_valid_headers = self.validate_headers()
            if not is_valid_headers:
                missing_headers = ', '.join(self.missing_columns)
                raise ValidationError(f'Missing headers: {missing_headers}')
            # open the output csv file
            self.csv_output_file_path = os.path.join(
                TEMP_OUTPUT_DIRECTORY,
                f'{str(self.request.uuid)}'
            ) + '.csv'
            # try to remove existing file if any
            self.remove_csv_output_file()
            # open the csv writer
            self.csv_output_file = open(self.csv_output_file_path, 'w')
            self.csv_output_writer = csv.writer(self.csv_output_file)
            # get output headers
            self.get_output_headers()
            self.csv_output_writer.writerow(self.output_headers)
            # process rows
            total_count, success_count, error_count = self.process_rows()
            if total_count == 0:
                raise ValidationError(
                    'You have uploaded empty spreadsheet, '
                    'please check again.'
                )
            is_success = True
            self.request.total_count = total_count
            self.request.success_count = success_count
            self.request.error_count = error_count
            # generate success notes
            if (
                self.request.success_count > 0 and
                self.request.error_count == 0
            ):
                if self.preview:
                    self.request.success_notes = (
                        f'{format_entities_count(self.request.success_count)} '
                        'can be updated.'
                    )
                else:
                    self.request.success_notes = (
                        f'{format_entities_count(self.request.success_count)} '
                        'have been updated.'
                    )
            if (
                self.request.success_count > 0 and
                self.request.error_count > 0
            ):
                if self.preview:
                    self.request.success_notes = (
                        f'{format_entities_count(self.request.success_count)} '
                        'can be updated. '
                        f'{format_entities_count(self.request.error_count)} '
                        'have errors.'
                    )
                else:
                    self.request.success_notes = (
                        f'{format_entities_count(self.request.success_count)} '
                        'have been updated. '
                        f'{format_entities_count(self.request.error_count)} '
                        'have errors.'
                    )
            self.request.save(update_fields=[
                'total_count', 'success_count', 'error_count', 'success_notes'
            ])
        except Exception as ex:
            logger.error(ex)
            logger.error(traceback.format_exc())
            error = str(ex)
            is_success = False
        finally:
            self.process_ended(is_success, error)

    def read_headers(self):
        """Read column headers."""
        pass

    def process_headers(self):
        """Read headers."""
        if len(self.headers) == 0:
            raise RuntimeError('Headers are empty!')
        # find index of ucode_field in headers
        for idx, header in enumerate(self.headers):
            if header == self.request.ucode_field:
                self.ucode_index = idx
            check_id = (
                [id_field for id_field in self.request.id_fields if
                 id_field['field'] == header]
            )
            if len(check_id) > 0:
                self.id_fields[idx] = check_id[0]
            check_name = (
                [name_field for name_field in self.request.name_fields if
                 name_field['field'] == header]
            )
            if len(check_name) > 0:
                self.name_fields[idx] = check_name[0]

    def validate_headers(self):
        """Validate mandatory columns."""
        self.missing_columns = []
        if self.ucode_index == -1:
            self.missing_columns.append(self.request.ucode_field)
        for id_field in self.request.id_fields:
            field_name = id_field['field']
            check_fields = (
                [field for field in self.id_fields.values() if
                 field['field'] == field_name]
            )
            if len(check_fields) == 0:
                self.missing_columns.append(field_name)
        for name_field in self.request.name_fields:
            field_name = name_field['field']
            check_fields = (
                [field for field in self.name_fields.values() if
                 field['field'] == field_name]
            )
            if len(check_fields) == 0:
                self.missing_columns.append(field_name)
        return len(self.missing_columns) == 0

    def process_rows(self):
        """Base process rows."""
        total_count = 0
        success_count = 0
        error_count = 0
        return total_count, success_count, error_count

    def get_output_headers(self):
        """Set output headers."""
        self.output_headers = [
            'Country',
            'Level',
            'Ucode',
            'Default Name',
            'Default Code'
        ]
        self.id_field_idx_start = len(self.output_headers)
        # add id fields
        for _, id_field in self.id_fields.items():
            field_name = id_field['field']
            self.output_headers.append(field_name)
        self.name_field_idx_start = len(self.output_headers)
        # add name fields
        for _, name_field in self.name_fields.items():
            field_name = name_field['field']
            self.output_headers.append(field_name)
        self.status_field_idx_start = len(self.output_headers)
        # add status and errors
        self.output_headers.append('Status')
        self.output_headers.append('Errors')

    def get_default_output_row(self, ucode):
        output_row = [
            '-' for i in range(len(self.output_headers))
        ]
        output_row[2] = ucode
        return output_row

    def get_output_row_with_status(self, output_row, is_success, error):
        """Set output row with status and error."""
        # status column should the last 2 index
        success_text = 'OK' if self.preview else 'SUCCESS'
        output_row[-2] = success_text if is_success else 'ERROR'
        if error:
            output_row[-1] = error
        return output_row

    def update_output_row_with_entity(self, output_row,
                                      entity: GeographicalEntity):
        """Update output row with entity information."""
        output_row[0] = entity.ancestor.label if entity.ancestor else ''
        if entity.level == 0:
            output_row[0] = entity.label
        output_row[1] = entity.level
        output_row[3] = entity.label
        output_row[4] = entity.internal_code
        # return the output row and starting idx for id/name field
        return output_row, 5

    def process_row(self, row):
        """
        Process individual row.

        :param row: list of cell values
        :return: Tuple of is_success, updated_row
        """
        ucode = row[self.ucode_index]
        updated_row = self.get_default_output_row(ucode)
        try:
            # parse ucode
            unique_code, version_number = parse_unique_code(ucode)
        except ValueError as ex:
            return (
                False,
                self.get_output_row_with_status(updated_row, False, str(ex))
            )
        # find entity
        entity = GeographicalEntity.objects.select_related(
            'ancestor'
        ).defer(
            'geometry', 'ancestor__geometry'
        ).filter(
            dataset=self.request.dataset,
            unique_code=unique_code,
            unique_code_version=version_number,
            is_approved=True
        )
        if not entity.exists():
            return (
                False,
                self.get_output_row_with_status(
                    updated_row, False, f'Entity {ucode} does not exist')
            )
        entity = entity.first()
        updated_row, new_idx_start = self.update_output_row_with_entity(
            updated_row, entity)
        # fetch entity metadata: existing names, existing ids
        entity_ids = list(EntityId.objects.select_related('code').filter(
            geographical_entity=entity
        ).all())
        entity_names = list(EntityName.objects.filter(
            geographical_entity=entity
        ).all())
        # process new id_fields
        id_errors = []
        new_ids = []
        updated_ids = []
        for id_field_idx, id_field in self.id_fields.items():
            field_name = id_field['field']
            id_value = self.row_value(row, id_field_idx)
            # update output row
            updated_row[new_idx_start] = id_value
            new_idx_start += 1
            # validate non-empty value
            if id_value == '':
                continue
            # get the id_type
            id_type_in = id_field['idType']
            id_type = (
                [id for id in self.id_types if id.id == id_type_in['id']]
            )
            if len(id_type) == 0:
                id_errors.append(f'Invalid id type of {field_name}')
                continue
            id_type = id_type[0]
            # validate non-duplicate id type
            existing_ids = (
                [id for id in entity_ids if id.code.id == id_type.id]
            )
            if len(existing_ids) > 0:
                # if there is existing id, overwrite if it's not default id
                existing_id = existing_ids[0]
                if existing_id.default:
                    id_errors.append(
                        'Cannot overwrite default ID '
                        f'{existing_id.code.name}')
                else:
                    existing_id.value = id_value
                    updated_ids.append(existing_id)
            else:
                # add to new_ids
                new_ids.append(EntityId(
                    geographical_entity=entity,
                    code=id_type,
                    value=id_value
                ))

        # get max value of idx in EntityName
        name_idx_start = 0
        max_idx_res = EntityName.objects.filter(
            geographical_entity=entity
        ).aggregate(Max('idx'))
        if max_idx_res:
            name_idx_start = (
                max_idx_res['idx__max'] + 1 if
                max_idx_res['idx__max'] is not None else 0
            )
        # process new name_fields
        name_errors = []
        new_names = []
        updated_names = []
        for name_field_idx, name_field in self.name_fields.items():
            field_name = name_field['field']
            name_value = self.row_value(row, name_field_idx)
            # update output row
            updated_row[new_idx_start] = name_value
            new_idx_start += 1
            # validate non-empty value
            if name_value == '':
                continue
            # get the language
            lang_in = name_field['selectedLanguage']
            language = [lang for lang in self.languages if lang.id == lang_in]
            if len(language) == 0:
                name_errors.append(f'Invalid language of {field_name}')
                continue
            language = language[0]
            name_label = name_field['label']
            if name_label:
                # find duplicate by label
                existing_names = (
                    [name for name in entity_names
                     if name.label == name_label]
                )
                if len(existing_names) > 0:
                    existing_name = existing_names[0]
                    existing_name.name = name_value
                    existing_name.language = language
                    updated_names.append(existing_name)
                    continue
            else:
                # find duplicate name by language and value
                is_existing_name = (
                    [name for name in entity_names if
                     name.name == name_value and
                     name.language.id == lang_in and
                     not name.label]
                )
                if len(is_existing_name) > 0:
                    name_errors.append(
                        f'Existing name {name_value} with '
                        f'language {language.name}')
                    continue
            # add to new_names
            new_names.append(EntityName(
                name=name_value,
                geographical_entity=entity,
                language=language,
                label=name_label,
                idx=name_idx_start
            ))
            name_idx_start += 1

        # insert record for new ids and new names
        if not self.preview:
            if new_ids:
                EntityId.objects.bulk_create(new_ids)
            if updated_ids:
                EntityId.objects.bulk_update(updated_ids, ['value'])
            if new_names:
                EntityName.objects.bulk_create(new_names)
            if updated_names:
                EntityName.objects.bulk_update(
                    updated_names, ['name', 'language'])
        # success if there is new ids/names are inserted or
        # existing ids are updated
        success = (
            len(new_ids) > 0 or len(new_names) > 0 or len(updated_ids) > 0
        )
        errors = []
        if id_errors:
            errors.append('; '.join(id_errors))
        elif len(self.id_fields) > 0 and len(new_ids) == 0:
            # no imported id_fields
            errors.append('No valid ID value')
        if name_errors:
            errors.append('; '.join(name_errors))
        elif len(self.name_fields) > 0 and len(new_names) == 0:
            # no imported id_fields
            errors.append('No valid name value')
        if success:
            if entity.id not in self.entity_ids:
                self.entity_ids.append(entity.id)
        return success, self.get_output_row_with_status(
            updated_row, success, '; '.join(errors))

    def row_value(self, row, index):
        """
        Get row value by index
        :param row: row data
        :param index: index
        :return: row value
        """
        row_value = ''
        try:
            row_value = row[index]
            row_value = row_value.replace('\xa0', ' ')
            row_value = row_value.replace('\xc2', '')
            row_value = row_value.replace('\\xa0', '')
            row_value = row_value.strip()
            row_value = re.sub(' +', ' ', row_value)
        except KeyError:
            pass
        return row_value

    def validate_input_file(self):
        """Validate input file from BatchEntityEdit obj."""
        self.read_headers()
        if self.total_rows < 1:
            return False, (
                'You have uploaded empty spreadsheet, '
                'please check again.'
            )
        # check if has headers
        if len(self.headers) < 2:
            return False, (
                'You have uploaded spreadsheet with invalid headers, '
                'the headers must include ucode field and '
                'at least one ID/Name field, '
                'please check again.'
            )
        # store headers and total count
        self.request.total_count = self.total_rows
        self.request.headers = self.headers
        self.request.save(update_fields=['total_count', 'headers'])
        return True, None

    def store_entity_edit_result(self, row_idx, output_row):
        new_codes = (
            output_row[self.id_field_idx_start:self.name_field_idx_start] if
            self.name_field_idx_start > self.id_field_idx_start else []
        )
        new_names = (
            output_row[
                self.name_field_idx_start:self.status_field_idx_start] if
            self.status_field_idx_start > self.name_field_idx_start else []
        )
        if self.preview:
            self.entity_results.append(
                EntityEditResult(
                    batch_edit=self.request,
                    row_idx=row_idx,
                    ucode=output_row[2],
                    level=output_row[1],
                    country=output_row[0],
                    default_name=output_row[3],
                    default_code=output_row[4],
                    status=output_row[self.status_field_idx_start],
                    errors=output_row[self.status_field_idx_start + 1],
                    new_names=new_names,
                    new_codes=new_codes
                )
            )
        else:
            result = EntityEditResult.objects.filter(
                batch_edit=self.request,
                row_idx=row_idx,
            ).first()
            if result:
                result.ucode = output_row[2]
                result.level = output_row[1]
                result.country = output_row[0]
                result.default_name = output_row[3]
                result.default_code = output_row[4]
                result.status = output_row[self.status_field_idx_start]
                result.errors = output_row[self.status_field_idx_start + 1]
                result.new_names = new_names
                result.new_codes = new_codes
                self.entity_results.append(result)

    def execute_batch_insert_edit_result(self, end=False):
        should_execute = end or len(self.entity_results) % 250 == 0
        if not should_execute:
            return
        if self.preview:
            # use create
            EntityEditResult.objects.bulk_create(self.entity_results)
        else:
            # use update
            EntityEditResult.objects.bulk_update(
                self.entity_results,
                fields=[
                    'ucode',
                    'level',
                    'country',
                    'default_name',
                    'default_code',
                    'status',
                    'errors',
                    'new_names',
                    'new_codes'
                ]
            )
        self.entity_results.clear()


class CSVBatchEntityEditImporter(BatchEntityEditBaseImporter):

    def read_headers(self):
        with self.request.input_file.open('rb') as csv_file:
            file = csv_file.read().decode(
                'utf-8', errors='ignore').splitlines()
            csv_reader = csv.reader(file)
            self.headers = next(csv_reader)
            self.total_rows = sum(1 for row in csv_reader)

    def process_rows(self):
        success_count = 0
        error_count = 0
        line_count = 0
        with self.request.input_file.open('rb') as csv_file:
            file = csv_file.read().decode(
                'utf-8', errors='ignore').splitlines()
            csv_reader = csv.reader(file)
            for row in csv_reader:
                if line_count == 0:
                    line_count += 1
                    continue
                else:
                    is_row_success, output_row = self.process_row(row)
                    if is_row_success:
                        success_count += 1
                    else:
                        error_count += 1
                    self.csv_output_writer.writerow(output_row)
                    self.store_entity_edit_result(line_count - 1, output_row)
                self.execute_batch_insert_edit_result()
                self.on_update_progress(line_count, self.total_rows)
                line_count += 1
        self.execute_batch_insert_edit_result(True)
        return line_count - 1, success_count, error_count


class ExcelBatchEntityEditImporter(BatchEntityEditBaseImporter):

    def read_headers(self):
        self.headers = []
        with self.request.input_file.open('rb') as excel_file:
            wb_obj = openpyxl.load_workbook(excel_file, data_only=True)
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            self.total_rows = sheet_obj.max_row - 1
            # Loop will print all columns name
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=1, column=i)
                self.headers.append(cell_obj.value)

    def process_rows(self):
        success_count = 0
        error_count = 0
        line_count = 0
        with self.request.input_file.open('rb') as excel_file:
            wb_obj = openpyxl.load_workbook(excel_file, data_only=True)
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            m_row = sheet_obj.max_row
            if m_row < 2:
                # contains only header or empty
                return line_count, success_count, error_count
            for row_idx in range(2, m_row + 1):
                row = []
                for col_idx in range(1, max_col + 1):
                    cell_obj = sheet_obj.cell(row=row_idx, column=col_idx)
                    cell_value = (
                        str(cell_obj.value) if
                        cell_obj.value is not None else ''
                    )
                    row.append(cell_value)
                is_row_success, output_row = self.process_row(row)
                if is_row_success:
                    success_count += 1
                else:
                    error_count += 1
                self.csv_output_writer.writerow(output_row)
                self.store_entity_edit_result(line_count, output_row)
                self.execute_batch_insert_edit_result()
                line_count += 1
                self.on_update_progress(line_count, self.total_rows)
        self.execute_batch_insert_edit_result(True)
        return line_count, success_count, error_count


def get_entity_edit_importer(
        obj: BatchEntityEdit, preview: bool) -> BatchEntityEditBaseImporter:
    if obj.input_file is None:
        raise RuntimeError('Batch session does not have input file uploaded!')
    file_name = obj.input_file.name
    if (
        file_name.endswith('.xlsx') or
        file_name.endswith('.xls')
    ):
        return ExcelBatchEntityEditImporter(obj, preview)
    elif file_name.endswith('.csv'):
        return CSVBatchEntityEditImporter(obj, preview)
    return None
